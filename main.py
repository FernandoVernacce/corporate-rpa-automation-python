
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference


def registrar_log(mensagem):
    with open("log_execucao.txt", "a", encoding="utf-8") as log:
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        log.write = (f"[{data_hora}], {mensagem}\n")

def contar_status_arquivo(caminho_arquivo):
    try:
        workbook = load_workbook(caminho_arquivo)
        sheet = workbook.active

        pagos = 0
        pendentes = 0
        cancelados = 0

        for row in range(2, sheet.max_row + 1):
            status = sheet.cell(row=row, column=2).value

            if status:
                status = status.lower()

                if status == "pago":
                    pagos += 1
                elif status == "pendente":
                    pendentes += 1
                elif status == "cancelado":
                    cancelados += 1
                else:
                    registrar_log(f"Status inválido encontrado em {caminho_arquivo}: {status}")

                registrar_log(f"Arquivo processado com sucesso: {caminho_arquivo}")
        
        return pagos, pendentes, cancelados

    except Exception as e:
        registrar_log(f"Erro ao processar {caminho_arquivo}: {e}")
        return 0, 0, 0
    
    
def processar_pasta(nome_pasta):
    if not os.path.exists(nome_pasta):
        registrar_log("Pasta não encontrada")
        raise FileNotFoundError("A pasta especificada não existe")
    
    total_pagos = 0
    total_pendentes = 0
    total_cancelados = 0

    for arquivo in os.listdir(nome_pasta):
        if arquivo.endswith(".xlsx"):
            caminho_completo = os.path.join(nome_pasta, arquivo)

            print(f"Processando: {arquivo}")
            registrar_log(f"Iniciando processamento:{arquivo}")

            pagos, pendentes, cancelados = contar_status_arquivo(caminho_completo)

            total_pagos += pagos
            total_pendentes += pendentes
            total_cancelados += cancelados
    
    return total_pagos, total_pendentes, total_cancelados

def criar_relatorio(pagos, pendentes, cancelados):
    novo_wb = Workbook()
    sheet = novo_wb.active
    sheet.title = "Resumo Geral"

    sheet.append(["Tipo", "Quantidade"])
    sheet.append(["Pagos", pagos])
    sheet.append(["Pendentes", pendentes])
    sheet.append(["Cancelados", cancelados])

    # Criando gráfico
    grafico = BarChart()
    grafico.title = "Resumo de Pagamentos"
    grafico.y_axis.title = "Quantidade"
    grafico.x_axis.title = "Status"

    dados = Reference(sheet, min_col=2, min_row=1, max_row=4)
    categorias = Reference(sheet, min_col=1, min_row=2, max_row=4)

    grafico.add_data(dados, titles_from_data=True)
    grafico.set_categories(categorias)

    sheet.add_chart(grafico, "D6")

    novo_wb.save("relatorio_geral.xlsx")
    registrar_log("Relatório geral gerado com sucesso.")


def main():
    try:
        pasta = 'clientes'

        pagos, pendentes, cancelados = processar_pasta(pasta)

        print("\nResumo Final:")
        print(f"Pagos: {pagos}")
        print(f"Pendentes: {pendentes}")
        print(f"Cancelados: {cancelados}")

        criar_relatorio(pagos, pendentes, cancelados)

        print("\nAutomação concluída com sucesso!")
        registrar_log("Execução finalizada com sucesso.")

    except Exception as e:
        print(f"Erro crítico: {e}")
        registrar_log(f"Erro crítico na execução: {e}")

if __name__ == "__main__":
    main()




